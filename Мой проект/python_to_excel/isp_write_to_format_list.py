from ast import NotIn
from fileinput import filename
from importlib.resources import path
from multiprocessing.sharedctypes import Value
import pandas as pd
import os
import openpyxl as ox
from pandas.io.excel import ExcelWriter
from openpyxl import Workbook
import numpy as np
import time
#pip install pyxlsb  # может понадобиться

#from timer import Timer
#t = Timer()
#t.start()
start = time.time()
cols = [i for i in range(0,19)] +[i for i in range(25,32)]+[i for i in range(41,72)]
#print (cols)
df = pd.read_excel("d:\ход работ\ГДДО22 СВОД.xlsb",sheet_name='ПЛАН', usecols=cols, index_col=False, header=4)
print (df.head())
isp=df['col31'].unique().tolist()
print(isp)



''''#  функция для очистки содержимого ячеек на листе книги не повреждая формат ячеек,
чистим только за размером записавыемого диапазона 100 строк
 или +100 строк от содержимого [A6:BU3000]
формат сохраняется.
'''

def clear_f ():    
    for i  in isp:
        xlsx_path = os.path.join(pathfile,"isp", "ГДДО22 СВОД_"+str(i.replace('"',"",4)) +".xlsx")
        wb = ox.load_workbook(filename=xlsx_path)
        #sheet = book['План']
        #sheet.clear("A4:BU3000")
        #book = ox.load_workbook('sample.xlsx') #get the file name
        sheet = wb.get_sheet_by_name('План') #get the sheet name
        len_i = len(df[df['col31']==i])
        
        for a in sheet['A'+str(len_i+4):'BE'+str(len_i+104)]: #you can set the range here 
            for cell in a:
                cell.value = None #set a value or null here
        wb.save(xlsx_path)  #book.save('sample.xlsx') #you can save it after if you like
        

''''# update_spreadsheet - функция для записи в книгу на существующий лист с ранее уст форматированием 
В видео несколько вариантов для перезаписи. Если использовать openpyxl напрямую, т.е. по схеме:
wb = ox.load_workbook(path)
wb[sheet_name].cell(startrow + ir, starcol + ic).value = _df.iloc[ir][ic]
то да, формат сохраняется.
'''
def update_spreadsheet(path : str, _df, starcol : int = 0, startrow : int = 5, sheetname : str = "План"):
    wb = ox.load_workbook(path)  # открываем файл
    for ir in range (0, len(_df)):
        for ic in range (0, len(_df.iloc[ir])):
            wb[sheetname].cell(startrow + ir, starcol + ic).value = _df.iloc[ir][ic]  # записываю данные в строки таблицы
    wb.save(path) 


# НАЧАЛО
# получаем путь к файлам
pathfile = 'd:\ход работ'

#clear_f()  # очистим ячейки файлов от старых данных + 100 строк от размера фрейма 

# ЗАПИШЕМ ДАННЫЕ в поименованные книги в лист "План", начиная с 6 строки и с 1 столбца
#df1 = df[df['col31'] =='ООО "Газпром трансгаз Югорск"']
ITC = ['АО "Газпром диагностика" ИТЦ "Ростов" участок Краснодар',
    'АО "Газпром диагностика" ИТЦ "Ростов"',\
    'АО "Газпром диагностика" ИТЦ "Санкт-Петербург"',\
    'АО "Газпром диагностика" ИТЦ "Санкт-Петербург" участок Приобье',\
    'АО "Газпром диагностика" ИТЦ "Санкт-Петербург" участок Томск',\
    'АО "Газпром диагностика" ИТЦ "Видное" (скважины)',\
    'АО "Газпром диагностика" ИТЦ "Видное" (отбраковка)',\
    'АО "Газпром диагностика" ИТЦ "Видное"']

ITC_ROS = ['АО "Газпром диагностика" ИТЦ "Ростов" участок Краснодар', 'АО "Газпром диагностика" ИТЦ "Ростов"']
ITC_SAN = ['АО "Газпром диагностика" ИТЦ "Санкт-Петербург"', 'АО "Газпром диагностика" ИТЦ "Санкт-Петербург" участок Приобье', 'АО "Газпром диагностика" ИТЦ "Санкт-Петербург" участок Томск']
ITC_VID = ['АО "Газпром диагностика" ИТЦ "Видное" (отбраковка)','АО "Газпром диагностика" ИТЦ "Видное"']


#for i in ITC_ROS:
df1 = df[df['col31'].isin(ITC_ROS)]
xlsx_path = os.path.join(pathfile,"isp", "ГДДО22 СВОД_АО Газпром диагностика ИТЦ Ростов.xlsx") 
update_spreadsheet(xlsx_path, df1.sample(len(df1)), sheetname="План", starcol=1, startrow=6)

#for i in ITC_SAN:
df2 = df[df['col31'].isin(ITC_SAN)]
xlsx_path = os.path.join(pathfile,"isp", "ГДДО22 СВОД_АО Газпром диагностика ИТЦ Санкт-Петербург.xlsx")  
update_spreadsheet(xlsx_path, df2.sample(len(df2)), sheetname="План", starcol=1, startrow=6)

#for i in ITC_VID:
df3 = df[df['col31'].isin(ITC_VID)]
xlsx_path = os.path.join(pathfile,"isp", "ГДДО22 СВОД_АО Газпром диагностика ИТЦ Видное.xlsx")
update_spreadsheet(xlsx_path, df3.sample(len(df3)), sheetname="План", starcol=1, startrow=6)

po = list(set(isp)- set(ITC))
print(po)
for i in po :    
    xlsx_path = os.path.join(pathfile,"isp", "ГДДО22 СВОД_"+str(i.replace('"',"",4)) +".xlsx")
    df1 = df[df['col31']==i]
    update_spreadsheet(xlsx_path, df1.sample(len(df1)), sheetname="План", starcol=1, startrow=6)
    
print('done')
end = time.time()
print(format(end-start))


