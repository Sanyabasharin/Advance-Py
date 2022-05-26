from fileinput import filename
from importlib.resources import path
from multiprocessing.sharedctypes import Value
import pandas as pd
import os
import openpyxl as ox
from pandas.io.excel import ExcelWriter
from openpyxl import Workbook
import numpy as np
import time
#pip install pyxlsb  # может понадобиться

#from timer import Timer
#t = Timer()
#t.start()
start = time.time()
cols = [i for i in range(0,19)] +[i for i in range(25,32)]+[i for i in range(41,72)]
#print (cols)
df = pd.read_excel("d:\ход работ\ГДДО22 СВОД.xlsb",sheet_name='ПЛАН', usecols=cols, index_col=False, header=4)
print (df.head())
deo=df['col2'].unique().tolist()
print(deo)



''''#  функция для очистки содержимого ячеек на листе книги не повреждая формат ячеек,
чистим только за размером записавыемого диапазона 100 строк
 или +100 строк от содержимого [A6:BU3000]
формат сохраняется.
'''

def clear_f ():    
    for i in deo:
        xlsx_path = os.path.join(pathfile,"deo", "ГДДО22 СВОД_"+str(i.replace('"',"",2)) +".xlsx")
        wb = ox.load_workbook(filename=xlsx_path)
        #sheet = book['План']
        #sheet.clear("A4:BU3000")
        #book = ox.load_workbook('sample.xlsx') #get the file name
        sheet = wb.get_sheet_by_name('План') #get the sheet name
        len_i = len(df[df['col2']==i])
        
        for a in sheet['A'+str(len_i+4):'BE'+str(len_i+104)]: #you can set the range here 
            for cell in a:
                cell.value = None #set a value or null here
        wb.save(xlsx_path)  #book.save('sample.xlsx') #you can save it after if you like
        

''''# update_spreadsheet - функция для записи в книгу на существующий лист с ранее уст форматированием 
В видео несколько вариантов для перезаписи. Если использовать openpyxl напрямую, т.е. по схеме:
wb = ox.load_workbook(path)
wb[sheet_name].cell(startrow + ir, starcol + ic).value = _df.iloc[ir][ic]
то да, формат сохраняется.
'''
def update_spreadsheet(path : str, _df, starcol : int = 0, startrow : int = 5, sheetname : str = "План"):
    wb = ox.load_workbook(path)  # открываем файл
    for ir in range (0, len(_df)):
        for ic in range (0, len(_df.iloc[ir])):
            wb[sheetname].cell(startrow + ir, starcol + ic).value = _df.iloc[ir][ic]  # записываю данные в строки таблицы
    wb.save(path) 


# НАЧАЛО
# получаем путь к файлам
pathfile = 'd:\ход работ'

clear_f()  # очистим ячейки файлов от старых данных + 100 строк от размера фрейма 

# ЗАПИШЕМ ДАННЫЕ в поименованные книги в лист "План", начиная с 6 строки и с 1 столбца
#df1 = df[df['col2'] =='ООО "Газпром трансгаз Югорск"']
for i in deo:
    xlsx_path = os.path.join(pathfile,"deo", "ГДДО22 СВОД_"+str(i.replace('"',"",2)) +".xlsx")
    df1 = df[df['col2']==i]  
    update_spreadsheet(xlsx_path, df1.sample(len(df1)), sheetname="План", starcol=1, startrow=6)

print('done')
end = time.time()
print(format(end-start))


